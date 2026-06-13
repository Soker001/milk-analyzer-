"""
DairyDash E2E Test Report Generator
Runs all 130 Selenium tests, collects results, writes professional .xlsx report
"""

import sys
import os
import time
import unittest
import importlib
from datetime import datetime

# Windows UTF-8 safe printing
def safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        text = " ".join(str(a).encode('ascii', errors='replace').decode('ascii') for a in args)
        print(text, **{k: v for k, v in kwargs.items() if k != 'end'})

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
CLR = {
    "header_bg":    "1A237E",   # Dark navy
    "header_fg":    "FFFFFF",
    "pass_bg":      "C8E6C9",   # Light green
    "pass_fg":      "1B5E20",
    "fail_bg":      "FFCDD2",   # Light red
    "fail_fg":      "B71C1C",
    "error_bg":     "FFE0B2",   # Light orange
    "error_fg":     "E65100",
    "skip_bg":      "E0E0E0",
    "skip_fg":      "424242",
    "section_bg":   "283593",   # Section header blue
    "alt_row_bg":   "E8EAF6",   # Alternating row
    "title_bg":     "0D47A1",
    "summary_bg":   "1565C0",
    "summary_fg":   "FFFFFF",
    "border_col":   "90A4AE",
}

STATUS_COLORS = {
    "PASS":  (CLR["pass_bg"],  CLR["pass_fg"]),
    "FAIL":  (CLR["fail_bg"],  CLR["fail_fg"]),
    "ERROR": (CLR["error_bg"], CLR["error_fg"]),
    "SKIP":  (CLR["skip_bg"],  CLR["skip_fg"]),
}


def thin_border(color="90A4AE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def cell_style(ws, row, col, value, bold=False, bg=None, fg="000000",
               align="left", wrap=False, size=10, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = thin_border()
    return c


def run_tests():
    """Import & run the test suite, return collected results."""
    # Ensure project root (parent of tests/) is on sys.path
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)

    # Reload to get fresh state
    for key in list(sys.modules.keys()):
        if "test_dairydash" in key:
            del sys.modules[key]

    from tests import test_dairydash as td

    loader  = unittest.TestLoader()
    suite   = loader.loadTestsFromModule(td)
    result  = td.TestResultCollector()

    safe_print("\n" + "="*70)
    safe_print(" DairyDash E2E Test Suite -- Running 130 Test Cases")
    safe_print("="*70)
    start = time.time()
    suite.run(result)
    elapsed = time.time() - start
    safe_print(f"\n[OK] Tests complete in {elapsed:.1f}s")
    return td.test_results, result, elapsed


def generate_xlsx(results, result_obj, elapsed):
    """Build a professional .xlsx report from the results list."""
    ts  = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"E2E_Test_Report_DairyDash_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ── SHEET 1: Cover / Summary ──────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Summary"
    ws_cover.sheet_view.showGridLines = False

    # Title block
    ws_cover.merge_cells("A1:H1")
    t = ws_cover["A1"]
    t.value     = "DairyDash -- Milk Analyzer & Smart Pricing System"
    t.font      = Font(bold=True, size=20, color="FFFFFF", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 45

    ws_cover.merge_cells("A2:H2")
    s = ws_cover["A2"]
    s.value     = "End-to-End Selenium Automation Test Report"
    s.font      = Font(bold=False, size=13, color="FFFFFF", name="Calibri")
    s.fill      = PatternFill("solid", fgColor=CLR["summary_bg"])
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[2].height = 28

    ws_cover.merge_cells("A3:H3")
    ws_cover["A3"].value = ""

    # Meta info
    meta = [
        ("Project",        "DairyDash — Milk Analyzer & Smart Pricing System"),
        ("Test Framework",  "Python Selenium 4 + webdriver-manager"),
        ("Browser",         "Google Chrome (Headless)"),
        ("Base URL",        "http://localhost:3000"),
        ("Test Date",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Duration",  f"{elapsed:.1f} seconds"),
        ("Prepared By",     "Automated QA Pipeline"),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        ws_cover.merge_cells(f"B{i}:D{i}")
        ws_cover.merge_cells(f"E{i}:H{i}")
        cell_style(ws_cover, i, 2, k, bold=True, bg="E3F2FD", size=11)
        cell_style(ws_cover, i, 5, v, size=11)
        ws_cover.row_dimensions[i].height = 22

    # Summary stats
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    errors  = sum(1 for r in results if r["status"] == "ERROR")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pct     = round(passed / total * 100, 1) if total else 0

    stats_row = 4 + len(meta) + 1
    ws_cover.merge_cells(f"A{stats_row}:H{stats_row}")
    h = ws_cover[f"A{stats_row}"]
    h.value     = "TEST EXECUTION SUMMARY"
    h.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    h.fill      = PatternFill("solid", fgColor=CLR["section_bg"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[stats_row].height = 26

    stats = [
        ("Total Test Cases",  total,   "1565C0", "FFFFFF"),
        ("[PASS] Passed",     passed,  "2E7D32", "FFFFFF"),
        ("[FAIL] Failed",     failed,  "C62828", "FFFFFF"),
        ("[ERR] Errors",      errors,  "E65100", "FFFFFF"),
        ("[SKIP] Skipped",    skipped, "546E7A", "FFFFFF"),
        ("Pass Rate",          f"{pct}%","004D40","FFFFFF"),
    ]
    sr = stats_row + 1
    for i, (label, val, bg, fg) in enumerate(stats):
        col = i + 2
        cell_style(ws_cover, sr,   col, label, bold=True, bg=bg,    fg=fg, align="center", size=11)
        cell_style(ws_cover, sr+1, col, val,   bold=True, bg=bg+"22" if len(bg)==6 else bg,
                   fg=bg, align="center", size=14)
        ws_cover.column_dimensions[get_column_letter(col)].width = 18
    ws_cover.row_dimensions[sr].height   = 30
    ws_cover.row_dimensions[sr+1].height = 36
    ws_cover.column_dimensions["A"].width = 4

    # Set column widths for meta
    for col in ["B","C","D","E","F","G","H"]:
        ws_cover.column_dimensions[col].width = 20

    # ── SHEET 2: Module-wise Summary ──────────────────────────────────────
    ws_mod = wb.create_sheet("Module Summary")
    ws_mod.sheet_view.showGridLines = False

    ws_mod.merge_cells("A1:F1")
    m = ws_mod["A1"]
    m.value     = "Module-wise Test Results"
    m.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    m.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    m.alignment = Alignment(horizontal="center", vertical="center")
    ws_mod.row_dimensions[1].height = 36

    headers_mod = ["Module", "Total", "Passed", "Failed", "Errors", "Pass %"]
    for c, h in enumerate(headers_mod, 1):
        cell_style(ws_mod, 2, c, h, bold=True, bg=CLR["section_bg"],
                   fg="FFFFFF", align="center", size=11)

    modules = {}
    for r in results:
        m = r.get("module", "General")
        if m not in modules:
            modules[m] = {"total":0,"pass":0,"fail":0,"error":0}
        modules[m]["total"] += 1
        if r["status"] == "PASS":  modules[m]["pass"]  += 1
        elif r["status"] == "FAIL":  modules[m]["fail"]  += 1
        elif r["status"] == "ERROR": modules[m]["error"] += 1

    for i, (mod, s) in enumerate(modules.items(), start=3):
        bg = CLR["alt_row_bg"] if i % 2 == 0 else "FFFFFF"
        pct_mod = round(s["pass"] / s["total"] * 100, 1) if s["total"] else 0
        row_vals = [mod, s["total"], s["pass"], s["fail"], s["error"], f"{pct_mod}%"]
        for c, v in enumerate(row_vals, 1):
            cell_style(ws_mod, i, c, v, bg=bg, align="center" if c > 1 else "left")
        ws_mod.row_dimensions[i].height = 22

    col_widths_mod = [30, 10, 10, 10, 10, 12]
    for c, w in enumerate(col_widths_mod, 1):
        ws_mod.column_dimensions[get_column_letter(c)].width = w

    # ── SHEET 3: Full Test Results ─────────────────────────────────────────
    ws = wb.create_sheet("Test Results")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = "DairyDash — Complete E2E Test Case Results"
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    cols = ["TC ID", "Module", "Test Title", "Test Steps",
            "Expected Result", "Actual Result", "Status", "Duration", "Remarks"]
    for c, h in enumerate(cols, 1):
        cell_style(ws, 2, c, h, bold=True, bg=CLR["section_bg"],
                   fg="FFFFFF", align="center", size=11)
    ws.row_dimensions[2].height = 28

    col_widths = [10, 20, 35, 45, 40, 40, 10, 12, 20]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, r in enumerate(results, start=3):
        alt_bg = CLR["alt_row_bg"] if i % 2 == 0 else "FFFFFF"
        bg, fg = STATUS_COLORS.get(r["status"], (alt_bg, "000000"))

        row_data = [
            r.get("tc_id",    ""),
            r.get("module",   ""),
            r.get("title",    ""),
            r.get("steps",    ""),
            r.get("expected", ""),
            r.get("actual",   ""),
            r.get("status",   ""),
            "~1s",
            r.get("remarks",  ""),
        ]
        for c, val in enumerate(row_data, 1):
            if c == 7:  # Status column — colour by status
                cell_style(ws, i, c, val, bold=True, bg=bg, fg=fg,
                           align="center", size=10)
            elif c in [1]:
                cell_style(ws, i, c, val, bold=True, bg=alt_bg,
                           align="center", size=10)
            else:
                cell_style(ws, i, c, val, bg=alt_bg, wrap=True, size=10)
        ws.row_dimensions[i].height = 38

    # ── SHEET 4: Failed Tests Detail ──────────────────────────────────────
    ws_fail = wb.create_sheet("Failed Tests")
    ws_fail.sheet_view.showGridLines = False
    ws_fail.merge_cells("A1:F1")
    f = ws_fail["A1"]
    f.value     = "Failed & Errored Test Details"
    f.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    f.fill      = PatternFill("solid", fgColor="B71C1C")
    f.alignment = Alignment(horizontal="center", vertical="center")
    ws_fail.row_dimensions[1].height = 30

    fail_hdrs = ["TC ID","Module","Test Title","Expected","Actual (Error)","Status"]
    for c, h in enumerate(fail_hdrs, 1):
        cell_style(ws_fail, 2, c, h, bold=True, bg="C62828", fg="FFFFFF", align="center")

    failures = [r for r in results if r["status"] in ("FAIL","ERROR")]
    if failures:
        for i, r in enumerate(failures, 3):
            row_data = [r["tc_id"], r["module"], r["title"],
                        r["expected"], r["actual"], r["status"]]
            for c, val in enumerate(row_data, 1):
                bg, fg = STATUS_COLORS.get(r["status"], ("FFFFFF","000000"))
                cell_style(ws_fail, i, c, val, bg="FFCDD2" if r["status"]=="FAIL" else "FFE0B2",
                           wrap=True, size=10)
    else:
        ws_fail.merge_cells("A3:F3")
        cell_style(ws_fail, 3, 1, "🎉  All tests passed! No failures found.",
                   bold=True, bg="C8E6C9", fg="1B5E20", align="center", size=12)

    fail_widths = [10, 20, 35, 40, 40, 10]
    for c, w in enumerate(fail_widths, 1):
        ws_fail.column_dimensions[get_column_letter(c)].width = w

    wb.save(out)
    safe_print(f"\n[REPORT] Report saved: {out}")
    return out


def main():
    safe_print("[START] Starting DairyDash E2E Test Runner")
    safe_print("   Make sure the server is running on http://localhost:3000\n")

    results, result_obj, elapsed = run_tests()

    # Print quick summary
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] in ("FAIL","ERROR"))
    safe_print(f"\n{'='*50}")
    safe_print(f"  TOTAL:   {total}")
    safe_print(f"  PASSED:  {passed}")
    safe_print(f"  FAILED:  {failed}")
    safe_print(f"  PASS %:  {round(passed/total*100,1) if total else 0}%")
    safe_print(f"{'='*50}")

    out = generate_xlsx(results, result_obj, elapsed)
    safe_print(f"\n[DONE] Open the report at:\n   {out}")
    return out


if __name__ == "__main__":
    main()
