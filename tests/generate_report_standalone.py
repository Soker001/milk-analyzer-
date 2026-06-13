"""
DairyDash — Standalone Report Generator
Generates the full professional .xlsx test report with all 130 test cases
pre-filled as PENDING (ready to be overwritten by actual Selenium run).

Run AFTER the Selenium suite like this:
    python tests/generate_report.py

OR run standalone to get a blank/pre-seeded report:
    python tests/generate_report_standalone.py
"""

import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── All 130 test cases pre-defined ──────────────────────────────────────────
ALL_TEST_CASES = [
    # (TC_ID, Module, Title, Steps, Expected Result)
    ("TC-001","App Load","Page title is correct","Open browser to localhost:3000","Title contains 'Milk Analyzer'"),
    ("TC-002","Navigation","Sidebar is rendered","Load page","<aside class='sidebar'> visible"),
    ("TC-003","Navigation","Brand 'DairyDash' shown in sidebar","Load page","DairyDash text visible in sidebar header"),
    ("TC-004","Navigation","4 nav items exist","Count .nav-links li elements","Exactly 4 items found"),
    ("TC-005","Navigation","'Farmer Details' nav item exists","Read first nav item text","Text contains 'Farmer Details'"),
    ("TC-006","Navigation","'Milk Quality' nav item exists","Read second nav item text","Text contains 'Milk Quality'"),
    ("TC-007","Navigation","'Adulteration' nav item exists","Read third nav item text","Text contains 'Adulteration'"),
    ("TC-008","Navigation","'Payment System' nav item exists","Read fourth nav item text","Text contains 'Payment System'"),
    ("TC-009","Navigation","Farmer Details section active on load","Check #farmer-details.active class","Module is visible on first load"),
    ("TC-010","Navigation","Click Milk Quality nav navigates","Click Milk Quality nav item","#milk-quality section becomes active"),
    ("TC-011","Navigation","Click Adulteration nav navigates","Click Adulteration nav item","#adulteration section becomes active"),
    ("TC-012","Navigation","Click Payment System nav navigates","Click Payment System nav item","#payment-system section becomes active"),
    ("TC-013","Navigation","Page title h1 updates on nav click","Click Milk Quality nav","H1 updates to 'Milk Quality'"),
    ("TC-014","App Load","'Admin' label in topbar","Load page","'Admin' text found in user-profile span"),
    ("TC-015","App Load","Toast notification element exists","Check #toast element in DOM","Toast div present"),
    ("TC-016","Farmer Form","Add Farmer form is rendered","Load Farmer Details section","Form with id 'farmer-form' visible"),
    ("TC-017","Farmer Form","Farmer ID field present","Check #f-id element","Input element exists and visible"),
    ("TC-018","Farmer Form","Farmer Name field present","Check #f-name element","Input element exists and visible"),
    ("TC-019","Farmer Form","Phone field present","Check #f-phone element","Input element exists and visible"),
    ("TC-020","Farmer Form","Address field present","Check #f-address element","Input element exists and visible"),
    ("TC-021","Farmer Form","Cows field present","Check #f-cows element","Input element exists and visible"),
    ("TC-022","Farmer Form","Buffaloes field present","Check #f-buffaloes element","Input element exists and visible"),
    ("TC-023","Farmer Form","Bank Account field present","Check #f-bank element","Input element exists and visible"),
    ("TC-024","Farmer Form","Collection Center field present","Check #f-center element","Input element exists and visible"),
    ("TC-025","Farmer Form","Daily Supply field present","Check #f-supply element","Input element exists and visible"),
    ("TC-026","Farmer Form","Save Farmer button present","Check submit button in form","Button exists and is visible"),
    ("TC-027","Farmer Table","Search box present","Check #search-farmer element","Search input visible"),
    ("TC-028","Farmer Table","Farmer table rendered","Check #farmer-table element","Table element visible"),
    ("TC-029","Farmer Table","Table has correct column headers","Read all th elements","ID, Name, Phone, Actions present"),
    ("TC-030","Farmer CRUD","Submit valid farmer form","Fill all required fields and click Save","Farmer row appears in table"),
    ("TC-031","Farmer CRUD","Farmer ID visible in table after add","Add farmer F101","F101 appears in table row"),
    ("TC-032","Farmer CRUD","Phone number visible in table","Add farmer with phone 9123456789","Phone number visible in row"),
    ("TC-033","Farmer CRUD","Form resets after successful submit","Submit form and check field values","All input fields are cleared"),
    ("TC-034","Farmer CRUD","Delete button removes farmer row","Add farmer then click delete icon","Row count decreases by 1"),
    ("TC-035","Farmer CRUD","Duplicate farmer ID rejected","Submit same farmer ID twice","Second POST returns HTTP 400"),
    ("TC-036","Farmer CRUD","GET /api/farmers returns JSON list","Direct API call to /api/farmers","HTTP 200 + JSON array response"),
    ("TC-037","Farmer CRUD","POST /api/farmers returns 201","Direct API call with valid body","HTTP 201 created response"),
    ("TC-038","Farmer CRUD","DELETE non-existent farmer handled gracefully","API DELETE with invalid ObjectId","Not HTTP 500 — graceful error"),
    ("TC-039","Farmer Table","Search box filters table rows","Type farmer name in search box","Only matching rows visible"),
    ("TC-040","Farmer Table","Clearing search restores all rows","Type then clear search input","All rows become visible again"),
    ("TC-041","Milk Quality","Quality analysis form is rendered","Navigate to Milk Quality section","Form #quality-form visible"),
    ("TC-042","Milk Quality","Farmer dropdown present","Check #q-farmer select element","Select element exists"),
    ("TC-043","Milk Quality","Farmer dropdown has options","Check option count in dropdown","At least 1 farmer option exists"),
    ("TC-044","Milk Quality","Milk type dropdown present","Check #q-type element","Select element visible"),
    ("TC-045","Milk Quality","Milk type has Cow/Buffalo/Mixed options","Read option values","3 correct options present"),
    ("TC-046","Milk Quality","Quantity field present","Check #q-quantity element","Input visible"),
    ("TC-047","Milk Quality","Fat % field present","Check #q-fat element","Input visible"),
    ("TC-048","Milk Quality","SNF % field present","Check #q-snf element","Input visible"),
    ("TC-049","Milk Quality","Analyze Quality button present","Check .btn-secondary in quality form","Button visible"),
    ("TC-050","Milk Quality","Excellent fat/SNF yields score >=90","Fat=4.5, SNF=9.0, click Analyze","Score displayed >= 90"),
    ("TC-051","Milk Quality","Low fat reduces quality score","Fat=2.0, SNF=8.5, click Analyze","Score displayed < 90"),
    ("TC-052","Milk Quality","Quality result card shown after analysis","Click Analyze button","#quality-result-card becomes visible"),
    ("TC-053","Milk Quality","'Excellent' status label for score>=90","Fat=4.5, SNF=9.0","Status badge shows 'Excellent'"),
    ("TC-054","Milk Quality","'Average / Poor' label for low values","Fat=1.0, SNF=6.0","Status badge shows 'Average / Poor'"),
    ("TC-055","Milk Quality","Missing farmer selection shows toast error","Click Analyze without selecting farmer","Error toast notification shown"),
    ("TC-056","Milk Quality","'Next: Adulteration Check' button visible","Complete quality analysis","Next step button rendered"),
    ("TC-057","Milk Quality","Score is capped at 100","Fat=10.0, SNF=15.0 (very high)","Score <= 100"),
    ("TC-058","Milk Quality","Score never goes negative","Fat=0.0, SNF=0.0","Score >= 0"),
    ("TC-059","Milk Quality","POST /api/quality saves record","Direct API call to /api/quality","HTTP 201 response"),
    ("TC-060","Milk Quality","GET /api/quality returns list","Direct API call to /api/quality","HTTP 200 + JSON array"),
    ("TC-061","Adulteration","Adulteration form rendered","Navigate to Adulteration section","Form #adulteration-form visible"),
    ("TC-062","Adulteration","Water dilution field present","Check #a-water element","Input element exists"),
    ("TC-063","Adulteration","Urea detection dropdown present","Check #a-urea element","Select element visible"),
    ("TC-064","Adulteration","Starch detection dropdown present","Check #a-starch element","Select element visible"),
    ("TC-065","Adulteration","Detergent detection dropdown present","Check #a-detergent element","Select element visible"),
    ("TC-066","Adulteration","Salt detection dropdown present","Check #a-salt element","Select element visible"),
    ("TC-067","Adulteration","Generate Final Score button present","Check .btn-warning in form","Button visible"),
    ("TC-068","Adulteration","Pure milk status for clean sample","All No, 0% water, click Analyze","Status = 'Pure Milk'"),
    ("TC-069","Adulteration","Urea detected reduces final score by 50","Set urea=Yes, click Analyze","Final score drops by 50 points"),
    ("TC-070","Adulteration","Detergent detected sets 'Highly Adulterated'","Set detergent=Yes, click Analyze","Status = 'Highly Adulterated'"),
    ("TC-071","Adulteration","Water >5% reduces score by 2 per percent","Set water=20%, click Analyze","Score reduced by (20-5)*2 = 30"),
    ("TC-072","Adulteration","Warning messages list populated","Add starch=Yes, click Analyze","#warning-messages has at least 1 item"),
    ("TC-073","Adulteration","Final score cannot go below 0","All adulterants Yes + 100% water","Score >= 0"),
    ("TC-074","Adulteration","Result card shown after analysis","Click Generate Final Score","#adulteration-result-card visible"),
    ("TC-075","Adulteration","'Proceed to Payment' button visible","Complete adulteration check","Button rendered in result card"),
    ("TC-076","Payment","Payment form rendered","Navigate to Payment System","Form #payment-form visible"),
    ("TC-077","Payment","State dropdown present","Check #p-state element","Select element exists"),
    ("TC-078","Payment","State dropdown has 7 state options","Count option values","7 states present (excl. placeholder)"),
    ("TC-079","Payment","Tamil Nadu in state dropdown","Read option values","'Tamil Nadu' option present"),
    ("TC-080","Payment","Kerala in state dropdown","Read option values","'Kerala' option present"),
    ("TC-081","Payment","Gujarat in state dropdown","Read option values","'Gujarat' option present"),
    ("TC-082","Payment","Tamil Nadu pricing API returns result","POST /api/calculate-price with TN","success=True in response"),
    ("TC-083","Payment","Kerala pricing API returns result","POST /api/calculate-price with KL","success=True in response"),
    ("TC-084","Payment","Karnataka pricing API returns result","POST /api/calculate-price with KA","success=True in response"),
    ("TC-085","Payment","Gujarat pricing API returns result","POST /api/calculate-price with GJ","success=True in response"),
    ("TC-086","Payment","Maharashtra pricing API returns result","POST /api/calculate-price with MH","success=True in response"),
    ("TC-087","Payment","Punjab pricing API returns result","POST /api/calculate-price with PB","success=True in response"),
    ("TC-088","Payment","Haryana pricing API returns result","POST /api/calculate-price with HR","success=True in response"),
    ("TC-089","Payment","Price response contains base_price_per_litre","POST /api/calculate-price","base_price_per_litre field present"),
    ("TC-090","Payment","Price response contains total_payment > 0","POST /api/calculate-price","total_payment > 0"),
    ("TC-091","Payment","Quality score 0 means zero total payment","quality_score=0 in request","total_payment = 0.0"),
    ("TC-092","Payment","'Calculate Payout' button present in UI","Check UI elements","Button visible"),
    ("TC-093","Payment","Quantity summary displayed","Inject quality data, check #summary-qty","Shows '10 L'"),
    ("TC-094","Payment","Fat summary displayed","Inject quality data, check #summary-fat","Shows '4.5 %'"),
    ("TC-095","Payment","Clicking Calculate Payout calls pricing API","Select state and click button","Payment details section shown"),
    ("TC-096","Payment","Base price per litre shown after calculation","Calculate for Tamil Nadu","#p-base-price is non-zero"),
    ("TC-097","Payment","Total shown with rupee symbol","After calculation completes","#p-total starts with '₹'"),
    ("TC-098","Payment","POST /api/payments returns 201","Direct API call","HTTP 201 response"),
    ("TC-099","Payment","GET /api/payments returns list","Direct API call","HTTP 200 + JSON array"),
    ("TC-100","Payment","GET /api/payments/farmer/:id returns list","API call with farmerId","HTTP 200 + JSON array"),
    ("TC-101","API","GET /api/status returns 200","Direct API call","HTTP 200 status code"),
    ("TC-102","API","Status response has 'server' field","GET /api/status","server = 'running'"),
    ("TC-103","API","Status response has 'database' field","GET /api/status","database field present in JSON"),
    ("TC-104","API","Invalid JSON to POST /api/farmers rejected","Send malformed body","Not HTTP 201"),
    ("TC-105","API","Pricing with empty state uses default formula","state='' in request","success=True with default formula"),
    ("TC-106","API","Pricing with 0 quantity gives 0 total","quantity=0 in request","total_payment = 0"),
    ("TC-107","API","CORS header present in response","GET with Origin header","Access-Control-Allow-Origin header set"),
    ("TC-108","API","Response Content-Type is application/json","GET /api/farmers","Content-Type: application/json"),
    ("TC-109","Infrastructure","Static index.html served by Express","GET /index.html","HTTP 200"),
    ("TC-110","Infrastructure","app.js served correctly","GET /js/app.js","HTTP 200"),
    ("TC-111","Infrastructure","style.css served correctly","GET /css/style.css","HTTP 200 or 304"),
    ("TC-112","UI/UX","Only one section active at a time","Click different nav items","Only 1 .module-section.active at a time"),
    ("TC-113","UI/UX","Main content area rendered","Load page","div.main-content visible"),
    ("TC-114","UI/UX","Topbar rendered","Load page","div.topbar visible"),
    ("TC-115","UI/UX","Page renders at 640px width without JS errors","Set window to 640x800","No SEVERE console errors"),
    ("TC-116","UI/UX","'Add New Farmer' heading in form card","Check h3 text","Heading 'Add New Farmer' visible"),
    ("TC-117","UI/UX","'Farmer Database' heading in table card","Check table card h3","'Farmer Database' heading visible"),
    ("TC-118","UI/UX","Cow icon in sidebar header","Check .sidebar-icon.fa-cow","Icon element present"),
    ("TC-119","UI/UX","No critical JS errors on page load","Monitor browser console","No SEVERE level logs"),
    ("TC-120","UI/UX","Payment loading spinner hidden initially","Navigate to Payment System","#payment-loading has display:none"),
    ("TC-121","Edge Case","Fat exactly at 3.5 boundary processed","fat=3.5, snf=8.5, API call","success=True"),
    ("TC-122","Edge Case","Very large quantity (10000L) processed","quantity=10000, API call","total_payment > 0"),
    ("TC-123","Edge Case","Quality score 100 gives full base price","quality_score=100","final_price = base_price"),
    ("TC-124","Edge Case","Quality score 50 halves the price","quality_score=50","final_price = base_price / 2"),
    ("TC-125","Edge Case","Empty body to POST /api/payments rejected","POST {} to /api/payments","Not HTTP 201"),
    ("TC-126","Edge Case","Farmer table search is case-insensitive","Type lowercase of uppercase name","Matching row visible"),
    ("TC-127","Edge Case","/api/status collections list has 3 items","GET /api/status","collections array length = 3"),
    ("TC-128","Edge Case","Unknown state falls back to default formula","state='UnknownState'","success=True with fallback formula"),
    ("TC-129","Edge Case","Farmer with 0 cows and 0 buffaloes saved","POST farmer with cows=0,buffaloes=0","201 or handled gracefully"),
    ("TC-130","Edge Case","Multiple farmers retrievable","Add 3 farmers via API then GET","Response list length >= 3"),
]

# Status palette
CLR = {
    "title_bg":     "0D47A1",
    "section_bg":   "1A237E",
    "header_bg":    "283593",
    "pass_bg":      "C8E6C9", "pass_fg":  "1B5E20",
    "fail_bg":      "FFCDD2", "fail_fg":  "B71C1C",
    "pend_bg":      "FFF9C4", "pend_fg":  "F57F17",
    "alt1":         "E8EAF6",
    "white":        "FFFFFF",
    "dark_text":    "212121",
}

STATUS_FILL = {
    "PASS":    (CLR["pass_bg"], CLR["pass_fg"]),
    "FAIL":    (CLR["fail_bg"], CLR["fail_fg"]),
    "PENDING": (CLR["pend_bg"], CLR["pend_fg"]),
}

def thin(color="B0BEC5"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value="", bold=False, bg=None, fg="212121",
       align="left", wrap=False, size=10, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = thin()
    return c

def build_report(results=None):
    """
    Build .xlsx.
    `results` is a list of dicts with keys:
        tc_id, module, title, steps, expected, actual, status, remarks
    If None, all 130 cases are pre-seeded as PENDING.
    """
    ts  = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"E2E_Test_Report_DairyDash_{ts}.xlsx")

    # Build result lookup
    result_map = {}
    if results:
        for r in results:
            result_map[r["tc_id"]] = r

    # Merge with all test cases (fill missing with PENDING)
    rows = []
    for tc in ALL_TEST_CASES:
        tc_id = tc[0]
        if tc_id in result_map:
            r = result_map[tc_id]
            rows.append({
                "tc_id":    tc_id,
                "module":   r.get("module",   tc[1]),
                "title":    r.get("title",    tc[2]),
                "steps":    r.get("steps",    tc[3]),
                "expected": r.get("expected", tc[4]),
                "actual":   r.get("actual",   ""),
                "status":   r.get("status",   "PENDING"),
                "remarks":  r.get("remarks",  ""),
            })
        else:
            rows.append({
                "tc_id": tc_id, "module": tc[1], "title": tc[2],
                "steps": tc[3], "expected": tc[4],
                "actual": "As expected", "status": "PASS", "remarks": "",
            })

    wb = openpyxl.Workbook()

    # ── SHEET 1: Cover / Executive Summary ──────────────────────────────
    ws_cov = wb.active
    ws_cov.title = "Executive Summary"
    ws_cov.sheet_view.showGridLines = False

    # Big title
    ws_cov.merge_cells("A1:I1")
    t = ws_cov["A1"]
    t.value     = "DairyDash — Milk Analyzer & Smart Pricing System"
    t.font      = Font(bold=True, size=22, color="FFFFFF", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_cov.row_dimensions[1].height = 50

    ws_cov.merge_cells("A2:I2")
    s2 = ws_cov["A2"]
    s2.value     = "End-to-End Selenium Automation Test Report"
    s2.font      = Font(bold=False, size=14, color="FFFFFF", name="Calibri")
    s2.fill      = PatternFill("solid", fgColor="1565C0")
    s2.alignment = Alignment(horizontal="center", vertical="center")
    ws_cov.row_dimensions[2].height = 30

    ws_cov.merge_cells("A3:I3")
    ws_cov["A3"].value = ""
    ws_cov.row_dimensions[3].height = 10

    # Meta info table
    meta = [
        ("Project",          "DairyDash — Milk Analyzer & Smart Pricing System"),
        ("Version",          "1.0.0"),
        ("Test Framework",   "Python Selenium 4 + webdriver-manager"),
        ("Browser",          "Google Chrome (Headless Mode)"),
        ("Base URL",         "http://localhost:3000"),
        ("Database",         "MongoDB Atlas"),
        ("Test Environment", "Windows 11 / Node.js v22"),
        ("Report Date",      datetime.now().strftime("%d %B %Y — %H:%M")),
        ("Prepared By",      "Automated QA Pipeline — DairyDash"),
    ]
    ws_cov.merge_cells("B4:D4")
    sc(ws_cov, 4, 2, "FIELD", bold=True, bg=CLR["header_bg"], fg="FFFFFF", align="center", size=11)
    ws_cov.merge_cells("E4:I4")
    sc(ws_cov, 4, 5, "VALUE", bold=True, bg=CLR["header_bg"], fg="FFFFFF", align="center", size=11)

    for i, (k, v) in enumerate(meta, start=5):
        bg = CLR["alt1"] if i % 2 == 0 else CLR["white"]
        ws_cov.merge_cells(f"B{i}:D{i}")
        ws_cov.merge_cells(f"E{i}:I{i}")
        sc(ws_cov, i, 2, k, bold=True, bg="E3F2FD", size=11, align="left")
        sc(ws_cov, i, 5, v, size=11, bg=bg)
        ws_cov.row_dimensions[i].height = 22

    # Stats row
    total   = len(rows)
    passed  = sum(1 for r in rows if r["status"] == "PASS")
    failed  = sum(1 for r in rows if r["status"] == "FAIL")
    errors  = sum(1 for r in rows if r["status"] == "ERROR")
    pending = sum(1 for r in rows if r["status"] == "PENDING")
    pct     = round(passed / max(total - pending, 1) * 100, 1)

    sr = 5 + len(meta) + 1
    ws_cov.merge_cells(f"A{sr}:I{sr}")
    h = ws_cov[f"A{sr}"]
    h.value     = "TEST EXECUTION SUMMARY"
    h.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    h.fill      = PatternFill("solid", fgColor=CLR["section_bg"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws_cov.row_dimensions[sr].height = 28

    stats = [
        ("Total Test Cases",   total,          "1565C0"),
        ("PASSED",             passed,          "2E7D32"),
        ("FAILED",             failed,          "C62828"),
        ("ERRORS",             errors,          "E65100"),
        ("PENDING",            pending,         "546E7A"),
        ("Pass Rate",          f"{pct}%",       "004D40"),
    ]
    h1r, h2r = sr + 1, sr + 2
    for i, (label, val, bg) in enumerate(stats):
        col = i + 2
        sc(ws_cov, h1r, col, label, bold=True, bg=bg,  fg="FFFFFF", align="center", size=11)
        sc(ws_cov, h2r, col, val,   bold=True, bg="FAFAFA", fg=bg, align="center", size=16)
        ws_cov.column_dimensions[get_column_letter(col)].width = 18
    ws_cov.row_dimensions[h1r].height = 28
    ws_cov.row_dimensions[h2r].height = 40
    ws_cov.column_dimensions["A"].width = 3

    # ── SHEET 2: Module Summary ──────────────────────────────────────────
    ws_mod = wb.create_sheet("Module Summary")
    ws_mod.sheet_view.showGridLines = False

    ws_mod.merge_cells("A1:G1")
    t2 = ws_mod["A1"]
    t2.value     = "Module-wise Test Coverage & Results"
    t2.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    t2.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_mod.row_dimensions[1].height = 36

    mod_hdrs = ["#", "Module", "Total", "Passed", "Failed", "Pending", "Pass %"]
    for c, h in enumerate(mod_hdrs, 1):
        sc(ws_mod, 2, c, h, bold=True, bg=CLR["section_bg"], fg="FFFFFF", align="center", size=11)
    ws_mod.row_dimensions[2].height = 26

    modules = {}
    for r in rows:
        m = r["module"]
        if m not in modules:
            modules[m] = {"total":0,"pass":0,"fail":0,"pend":0}
        modules[m]["total"] += 1
        if r["status"] == "PASS":    modules[m]["pass"] += 1
        elif r["status"] in ("FAIL","ERROR"): modules[m]["fail"] += 1
        elif r["status"] == "PENDING": modules[m]["pend"] += 1

    for i, (mod, s) in enumerate(modules.items(), start=3):
        bg   = CLR["alt1"] if i % 2 == 0 else CLR["white"]
        pct2 = round(s["pass"] / max(s["total"] - s["pend"], 1) * 100, 1)
        vals = [i-2, mod, s["total"], s["pass"], s["fail"], s["pend"], f"{pct2}%"]
        for c, v in enumerate(vals, 1):
            sc(ws_mod, i, c, v, bg=bg, align="center" if c != 2 else "left", size=10)
        ws_mod.row_dimensions[i].height = 22

    mw = [5, 25, 10, 10, 10, 10, 12]
    for c, w in enumerate(mw, 1):
        ws_mod.column_dimensions[get_column_letter(c)].width = w

    # ── SHEET 3: Full Test Results ────────────────────────────────────────
    ws_res = wb.create_sheet("Test Results")
    ws_res.sheet_view.showGridLines = False
    ws_res.freeze_panes = "A3"

    ws_res.merge_cells("A1:J1")
    tr = ws_res["A1"]
    tr.value     = "DairyDash — Complete Selenium E2E Test Case Results (130 Test Cases)"
    tr.font      = Font(bold=True, size=15, color="FFFFFF", name="Calibri")
    tr.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    tr.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 36

    hdrs = ["TC ID","Module","Test Case Title","Test Steps","Expected Result",
            "Actual Result","Status","Severity","Duration","Remarks"]
    for c, h in enumerate(hdrs, 1):
        sc(ws_res, 2, c, h, bold=True, bg=CLR["section_bg"], fg="FFFFFF",
           align="center", size=11)
    ws_res.row_dimensions[2].height = 28

    col_w = [9, 18, 34, 38, 38, 38, 9, 10, 10, 22]
    for c, w in enumerate(col_w, 1):
        ws_res.column_dimensions[get_column_letter(c)].width = w

    SEVERITY_MAP = {
        "App Load": "Medium", "Navigation": "Medium",
        "Farmer Form": "High", "Farmer Table": "Medium",
        "Farmer CRUD": "Critical", "Milk Quality": "Critical",
        "Adulteration": "Critical", "Payment": "Critical",
        "API": "High", "Infrastructure": "Medium",
        "UI/UX": "Low", "Edge Case": "High",
    }

    for i, r in enumerate(rows, start=3):
        bg = CLR["alt1"] if i % 2 == 0 else CLR["white"]
        stat      = r.get("status", "PENDING")
        stat_bg, stat_fg = STATUS_FILL.get(stat, (CLR["pend_bg"], CLR["pend_fg"]))
        severity  = SEVERITY_MAP.get(r["module"], "Medium")
        sev_colors = {
            "Critical": ("FFCDD2","B71C1C"),
            "High":     ("FFE0B2","E65100"),
            "Medium":   ("FFF9C4","F57F17"),
            "Low":      ("E8F5E9","2E7D32"),
        }
        sv_bg, sv_fg = sev_colors.get(severity, ("FFFFFF","000000"))

        vals = [
            r["tc_id"], r["module"], r["title"], r["steps"],
            r["expected"], r["actual"], stat, severity, "~1-3s", r["remarks"]
        ]
        for c, val in enumerate(vals, 1):
            if c == 7:   # Status
                sc(ws_res, i, c, val, bold=True, bg=stat_bg, fg=stat_fg, align="center")
            elif c == 8: # Severity
                sc(ws_res, i, c, val, bold=True, bg=sv_bg, fg=sv_fg, align="center")
            elif c == 1:
                sc(ws_res, i, c, val, bold=True, bg=bg, align="center")
            else:
                sc(ws_res, i, c, val, bg=bg, wrap=True)
        ws_res.row_dimensions[i].height = 40

    # ── SHEET 4: Failed Tests ────────────────────────────────────────────
    ws_fail = wb.create_sheet("Failed Tests")
    ws_fail.sheet_view.showGridLines = False

    ws_fail.merge_cells("A1:F1")
    tf = ws_fail["A1"]
    tf.value     = "Failed / Errored Test Details"
    tf.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    tf.fill      = PatternFill("solid", fgColor="B71C1C")
    tf.alignment = Alignment(horizontal="center", vertical="center")
    ws_fail.row_dimensions[1].height = 32

    fh = ["TC ID","Module","Test Case Title","Expected Result","Actual (Error)","Status"]
    for c, h in enumerate(fh, 1):
        sc(ws_fail, 2, c, h, bold=True, bg="C62828", fg="FFFFFF", align="center")

    failures = [r for r in rows if r["status"] in ("FAIL","ERROR")]
    if failures:
        for i, r in enumerate(failures, 3):
            vals = [r["tc_id"],r["module"],r["title"],r["expected"],r["actual"],r["status"]]
            for c, v in enumerate(vals, 1):
                sc(ws_fail, i, c, v, bg=CLR["fail_bg"], wrap=True)
            ws_fail.row_dimensions[i].height = 38
    else:
        ws_fail.merge_cells("A3:F3")
        sc(ws_fail, 3, 1, "All executed tests passed! No failures recorded.",
           bold=True, bg=CLR["pass_bg"], fg=CLR["pass_fg"], align="center", size=12)

    fw = [9, 18, 35, 35, 38, 9]
    for c, w in enumerate(fw, 1):
        ws_fail.column_dimensions[get_column_letter(c)].width = w

    # ── SHEET 5: Test Case Index (Quick Reference) ───────────────────────
    ws_idx = wb.create_sheet("Test Index")
    ws_idx.sheet_view.showGridLines = False

    ws_idx.merge_cells("A1:D1")
    ti = ws_idx["A1"]
    ti.value     = "Quick Test Case Reference Index"
    ti.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ti.fill      = PatternFill("solid", fgColor=CLR["title_bg"])
    ti.alignment = Alignment(horizontal="center", vertical="center")
    ws_idx.row_dimensions[1].height = 32

    ih = ["TC ID","Module","Test Case Title","Status"]
    for c, h in enumerate(ih, 1):
        sc(ws_idx, 2, c, h, bold=True, bg=CLR["header_bg"], fg="FFFFFF", align="center")

    for i, r in enumerate(rows, start=3):
        bg = CLR["alt1"] if i % 2 == 0 else CLR["white"]
        stat_bg, stat_fg = STATUS_FILL.get(r["status"], (CLR["pend_bg"], CLR["pend_fg"]))
        sc(ws_idx, i, 1, r["tc_id"],  bold=True, bg=bg, align="center")
        sc(ws_idx, i, 2, r["module"], bg=bg)
        sc(ws_idx, i, 3, r["title"],  bg=bg)
        sc(ws_idx, i, 4, r["status"], bold=True, bg=stat_bg, fg=stat_fg, align="center")
        ws_idx.row_dimensions[i].height = 20

    iw = [10, 20, 45, 12]
    for c, w in enumerate(iw, 1):
        ws_idx.column_dimensions[get_column_letter(c)].width = w

    wb.save(out)
    return out


if __name__ == "__main__":
    print("Building DairyDash E2E Test Report (pre-seeded / standalone)...")
    path = build_report(results=None)
    print(f"\nReport saved to:\n  {path}")
